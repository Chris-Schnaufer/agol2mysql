#!/usr/bin/python3
""" This script moves data from an ESRI EXCEL spreadsheet, or directly from ESRI, 
to a MySql database
"""

import os
import argparse
import sys
import logging
from datetime import datetime
from getpass import getpass
from typing import Optional
import openpyxl
from openpyxl import load_workbook
import mysql.connector
from arcgis.gis import GIS

import a2database
from a2database import A2Database

# Check if we have the geometry transformation module
try:
    from osgeo import ogr
    from osgeo import osr
except ModuleNotFoundError:
    GEOM_CAN_TRANSFORM = False
else:
    GEOM_CAN_TRANSFORM = True

# The name of our script
SCRIPT_NAME = os.path.basename(__file__)

# Default host name to connect to the database
DEFAULT_HOST_NAME = 'localhost'

# Default number of heading lines in the EXCEL file
DEFAULT_NUM_HEADER_LINES = 1

# Default row that has the column names
DEFAULT_COL_NAMES_ROW = 1

# The default primary key database name for the sheet data
DEFAULT_PRIMARY_KEY_NAME = 'globalid'

# Default EPSG code for points
DEFAULT_GEOM_EPSG = 4326

# Default name for logging output
DEFAULT_LOG_FILENAME = 'data_xfer_excel.out'

# AGOL URL
DEFAULT_ESRI_URL = 'https://uagis.maps.arcgis.com'
# The client ID used to connect to AGOL
DEFAULT_SURVEY123_CLIENT_ID ='7X8jzC2i59RTyBVp'
# The feature item ID
DEFAULT_FEATURE_ITEM_ID = '7c8545bda6094875a5bf518de7f62b16'

# Argparse-related definitions
# Declare the progam description
ARGPARSE_PROGRAM_DESC = 'Uploads data from an ESRI generated excel spreadsheet to the ' \
                        'MySQL database'
# Epilog to argparse arguments
ARGPARSE_EPILOG = 'Duplicates are avoided by checking if the data already exists in the database'
# Host name help
ARGPARSE_HOST_HELP = f'The database host to connect to (default={DEFAULT_HOST_NAME})'
# Name of the database to connect to
ARGPARSE_DATABASE_HELP = 'The database to connect to'
# User name help
ARGPARSE_USER_HELP = 'The username to connect to the database with'
# Password help
ARGPARSE_PASSWORD_HELP = 'Prompt for the password used to connect to the database'
# Declare the help text for the EXCEL filename parameter (for argparse)
ARGPARSE_EXCEL_FILE_HELP = 'Path to the EXCEL file to upload'
# Declare the help text for the force deletion flag
ARGPARSE_UPDATE_HELP = 'Update existing data with new values (default is to skip updates)'
# Declare the help for no headers
ARGPARSE_HEADER_HELP = 'Specify the number of lines to consider as headings ' \
                       f'(default {DEFAULT_NUM_HEADER_LINES} lines)'
# Declare the help for where the column names are
ARGPARSE_COL_NAMES_ROW_HELP = 'Specify the row that contains the column names ' \
                              f'(default row is {DEFAULT_COL_NAMES_ROW})'
# Help text for specifying all the column names for the excel file
ARGPARSE_COL_NAMES_HELP = 'Comma seperated list of column names (used when column names are not ' \
                          'specified in the file)'
ARRGPARSE_COL_NAME_MAP_HELP = 'Maps case-sensitive column name in an excel file to a different ' \
                            'name. Use when the DB column name is different than the spreadsheet ' \
                            'name. Format as <sheet name>:<col name>=<DB col name>. ' \
                            'Eg: Mysheet:col 1=other_col'
# Help text fo specifying the primary key to use
ARGPARSE_PRIMARY_KEY_HELP = 'The name of the primary key to use when checking if data is ' \
                            f'already in the database (default "{DEFAULT_PRIMARY_KEY_NAME}"). ' \
                            'Also needs to be a column in the spreadsheet'
# Help for specifying columns that are to be considered point columns
ARGPARSE_POINT_COLS_HELP = 'The names of the X and Y columns in the spreadsheet that represent ' \
                           'a point type when creating a schema ("<x name>,<y name>")'
# Help for specifying the EPSG code that geometry coordinates are in
ARGPARSE_GEOMETRY_EPSG_HELP = 'The EPSG code of the coordinate system for the geometric objects ' \
                           f'(default is {DEFAULT_GEOM_EPSG})'
# Help for specifying the default database connection EPSG code
ARGPARSE_DATABASE_EPSG_HELP = 'The EPSG code of the database geometry ' \
                           f'(default is {DEFAULT_GEOM_EPSG})'
# Help text for verbose flag
ARGPARSE_VERBOSE_HELP = 'Display additional information as the script is run'
# Help for the reset flag
ARGPARSE_RESET_HELP = 'Deletes the contents of the tables to the loaded data (destroys old data)'
ARGPARSE_LOG_FILENAME_HELP = 'Change the name of the file where logging gets saved. This is a ' \
                             'destructive overwrite of existing files. Default logging file is ' \
                             f'named {DEFAULT_LOG_FILENAME}'
# Map a table name to another name
ARGPARSE_MAP_TABLE_NAME_HELP = 'Case-sensitive map a table name to a new name (intended to be ' \
                               'used when a source table name changes). E.g. previous_name=new_name'
# The endpoint URL
ARGPARSE_ESRI_ENDPOINT_HELP = f'URL to connect to. Default: {DEFAULT_ESRI_URL}'
# The AGOL application to connect to
ARGPARSE_ESRI_CLIENT_ID_HELP = 'The ID of the client to connect to. Default: ' \
                               f'{DEFAULT_SURVEY123_CLIENT_ID}'
# The feature of interest to get the schema from
ARGPARSE_ESRI_FEATURE_ID_HELP = 'The ID of the feature to get the database schema from. ' \
                                f'Default: {DEFAULT_FEATURE_ITEM_ID}'
# Lowering the debug level to DEBUG
ARGPARSE_LOGGING_DEBUG_HELP = 'Increases the logging level to include debugging messages'


def get_arguments(logger: logging.Logger) -> tuple:
    """ Returns the data from the parsed command line arguments
    Arguments:
        logger: the logging instance to use
    Returns:
        A tuple consisting of a string containing the EXCEL file name to process, and
        a dict of the command line options
    Exceptions:
        A ValueError exception is raised if the filename is not specified
    Notes:
        If an error is found, the script will exit with a non-zero return code
    """
    parser = argparse.ArgumentParser(prog=SCRIPT_NAME,
                                     description=ARGPARSE_PROGRAM_DESC,
                                     epilog=ARGPARSE_EPILOG)
    parser.add_argument('-excel_file',
                        help=ARGPARSE_EXCEL_FILE_HELP)
    parser.add_argument('-o', '--host', default=DEFAULT_HOST_NAME,
                        help=ARGPARSE_HOST_HELP)
    parser.add_argument('-d', '--database', help=ARGPARSE_DATABASE_HELP)
    parser.add_argument('-u', '--user', help=ARGPARSE_USER_HELP)
    parser.add_argument('-f', '--force', action='store_true',
                        help=ARGPARSE_UPDATE_HELP)
    parser.add_argument('--verbose', action='store_true',
                        help=ARGPARSE_VERBOSE_HELP)
    parser.add_argument('-p', '--password', action='store_true',
                        help=ARGPARSE_PASSWORD_HELP)
    parser.add_argument('--header', type=int, default=DEFAULT_NUM_HEADER_LINES,
                        help=ARGPARSE_HEADER_HELP)
    parser.add_argument('--col_names_row', type=int, default=DEFAULT_COL_NAMES_ROW,
                        help=ARGPARSE_COL_NAMES_ROW_HELP)
    parser.add_argument('--col_names', help=ARGPARSE_COL_NAMES_HELP)
    parser.add_argument('--col_name_map', action='append',
                        help=ARRGPARSE_COL_NAME_MAP_HELP)
    parser.add_argument('--point_cols', help=ARGPARSE_POINT_COLS_HELP)
    parser.add_argument('--geometry_epsg', type=int, default=DEFAULT_GEOM_EPSG,
                        help=ARGPARSE_GEOMETRY_EPSG_HELP)
    parser.add_argument('--database_epsg', type=int, default=DEFAULT_GEOM_EPSG,
                        help=ARGPARSE_DATABASE_EPSG_HELP)
    parser.add_argument('-k', '--key_name', default=DEFAULT_PRIMARY_KEY_NAME,
                        help=ARGPARSE_PRIMARY_KEY_HELP)
    parser.add_argument('--reset', action='store_true', help=ARGPARSE_RESET_HELP)
    parser.add_argument('--log_filename', default=DEFAULT_LOG_FILENAME,
                        help=ARGPARSE_LOG_FILENAME_HELP)
    parser.add_argument('-m', '--map_name', action='append',
                        help=ARGPARSE_MAP_TABLE_NAME_HELP)
    parser.add_argument('-ee', '--esri_endpoint', default=DEFAULT_ESRI_URL,
                        help=ARGPARSE_ESRI_ENDPOINT_HELP)
    parser.add_argument('-ec', '--esri_client_id', default=DEFAULT_SURVEY123_CLIENT_ID,
                        help=ARGPARSE_ESRI_CLIENT_ID_HELP)
    parser.add_argument('-ef', '--esri_feature_id', default=DEFAULT_FEATURE_ITEM_ID,
                        help=ARGPARSE_ESRI_FEATURE_ID_HELP)
    parser.add_argument('--debug', help=ARGPARSE_LOGGING_DEBUG_HELP)
    args = parser.parse_args()

    # Find the EXCEL file and the password (which is allowed to be eliminated)
    excel_file = None
    if args.excel_file:
        excel_file = args.excel_file
        try:
            with open(excel_file, encoding="utf-8"):
                pass
        except FileNotFoundError:
            logger.error(f'Unable to open EXCEL file {excel_file}')
            sys.exit(11)

    # Create the table name map
    table_name_map = {}
    if args.map_name and len(args.map_name) > 0:
        for one_map in args.map_name:
            if not '=' in one_map:
                logger.error(f'Invalid table name mapping {one_map}')
                sys.exit(13)
            old_name,new_name = one_map.split('=', 1)
            table_name_map[old_name] = new_name

    cmd_opts = {'force': args.force,
                'verbose': args.verbose,
                'host': args.host,
                'database': args.database,
                'user': args.user,
                'password': args.password,
                'header_lines': args.header,
                'col_names_row': args.col_names_row,
                'col_names': tuple(one_name.strip() for one_name in args.col_names.split(',')) \
                                        if args.col_names else None,
                'col_name_map': tuple(args.col_name_map) if args.col_name_map else tuple(),
                'point_col_x': args.point_cols.split(',')[0] if args.point_cols else None,
                'point_col_y': args.point_cols.split(',')[1] if args.point_cols else None,
                'geometry_epsg': args.geometry_epsg,
                'database_epsg': args.database_epsg,
                'primary_key': args.key_name,
                'reset': args.reset,
                'log_filename': args.log_filename,
                'table_name_map': table_name_map if table_name_map else None,
                'esri_endpoint': args.esri_endpoint,
                'esri_client_id': args.esri_client_id,
                'esri_feature_id': args.esri_feature_id,
                'debug': args.debug
               }

    # Postprocess column name mapping if there are any
    if cmd_opts['col_name_map']:
        new_maps = []
        for one_map in cmd_opts['col_name_map']:
            if not ':' in one_map or not '=' in one_map:
                logger.errror('Improperly formed column name mapping')
                sys.exit(12)
            sheet, cols = one_map.split(':', 1)
            old_col, db_col = cols.split('=', 1)
            new_maps.append({'sheet_name':sheet, 'sheet_col': old_col, 'db_col': db_col})
        cmd_opts['col_name_map'] = tuple(new_maps)

    # Return the loaded JSON
    return excel_file, cmd_opts


def init_logging(filename: str, level: int=logging.INFO) -> logging.Logger:
    """Initializes the logging
    Arguments:
        filename: name of the file to save logging to
        level: the logging level to use
    Return:
        Returns the created logger instance
    """
    logger = logging.getLogger()
    logger.setLevel(level if level is not None else logging.INFO)

    # Create formatter
    formatter = logging.Formatter('%(levelname)s: %(message)s')

    # Console handler
    cur_handler = logging.StreamHandler()
    cur_handler.setFormatter(formatter)
    logger.addHandler(cur_handler)

    # Output file handler
    cur_handler = logging.FileHandler(filename, mode='w')
    cur_handler.setFormatter(formatter)
    logger.addHandler(cur_handler)

    return logger


def transform_points(from_epsg: int, to_epsg: int, values: tuple) -> list:
    """Returns the list with the last num_values converted to the new coordinate system
    Arguments:
        from_epsg: original EPSG code
        to_epsg: EPSG code to transform the points to
        values: the list of values to transform (assumes X1, Y1, X2, Y2, ...)
    Returns:
        The list of transformed values
    """
    if len(values) % 2 or len(values) < 1:
        raise ValueError('Unable to transform points, an even number of X,Y values pairs ' \
                         'are needed')

    # Make sure we can transform points
    if not GEOM_CAN_TRANSFORM:
        raise ValueError('Unable to transform points, supporting osgeo module is not installed')

    # Transform the points
    return_values = []
    idx = 0
    from_sr = osr.SpatialReference()
    from_sr.ImportFromEPSG(int(from_epsg))
    to_sr = osr.SpatialReference()
    to_sr.ImportFromEPSG(int(to_epsg))
    transform = osr.CreateCoordinateTransformation(from_sr, to_sr)
    while idx < len(values):
        cur_geom = ogr.CreateGeometryFromWkt(f'POINT({values[idx]} {values[idx+1]} {from_epsg})')
        cur_geom.Transform(transform)
        return_values.extend((cur_geom.GetX(), cur_geom.GetY()))
        idx = idx + 2

    return return_values


def transform_geom_cols(col_names: tuple, col_values: tuple, geom_col_info: dict, from_epsg: int, \
                        to_epsg: int) -> list:
    """Transforms geometry point to the specified coordinate system
    Arguments:
        col_name: the column names of the table
        col_values: the values associated with the column names
        geom_col_info: the geometry column information
        from_epsg: the EPSG code to tranform from
        to_epsg: the EPSG code to transform to
    Return:
        Returns a list of column values with the geometry values transformed
    """
    # Check if we can avoid the transformations
    if from_epsg == to_epsg:
        return list(col_values)
    if len(geom_col_info['sheet_cols']) % 2:
        raise ValueError('Invalid number of X,Y column name pairs specified')

    pt_indexes = []
    pt_values = []
    idx = 0
    while idx < len(geom_col_info['sheet_cols']):
        x_idx = col_names.index(geom_col_info['sheet_cols'][idx])
        y_idx = col_names.index(geom_col_info['sheet_cols'][idx+1])
        pt_indexes.append(x_idx)
        pt_indexes.append(y_idx)
        pt_values.append(col_values[x_idx])
        pt_values.append(col_values[y_idx])
        idx += 2

    new_pt_values = transform_points(from_epsg, to_epsg, pt_values)

    return_values = list(col_values)
    for idx, idx_val in enumerate(pt_indexes):
        return_values[idx_val] = new_pt_values[idx]

    return return_values


def db_get_fk_constraints(conn: A2Database, table_name: str, logger: logging.Logger,
                          verbose: bool=False) -> Optional[dict]:
    """Returns the foreign key constraints that point to the specified table, not the table's
       constraints
    Arguments:
        conn: the database connection
        table_name: table that the constraints point to
        logger: logging instance
        verbose: prints additional information when True (truthy)
    Returns:
        A dictionary of the constraints' definitions, or None if no constraints are
        found.
        The dictionary is keyed by the constraint's name and each entry has the following
        format. The referenced columns (ref_columns) are ordered by their oridinal value.
        {
           contraint_name: {
            'table_name': the table the constraint belongs to
            'table_schema': the schema of the table
            'table_column': the constrained column
            'constraint_schema': the schema the constraint belongs to
            'ref_table_schema': the schema of the table the constraint references
            'ref_table_name': the table the constraint references
            'ref_columns': [
                column name 1,
                column name 2,
                ...
            ]
           }
        }, {...}
    """
    constraints = {}

    query = 'SELECT TABLE_SCHEMA,TABLE_NAME,COLUMN_NAME,CONSTRAINT_SCHEMA,CONSTRAINT_NAME,' \
            'REFERENCED_TABLE_SCHEMA,REFERENCED_TABLE_NAME,REFERENCED_COLUMN_NAME, '\
            'ORDINAL_POSITION FROM INFORMATION_SCHEMA.KEY_COLUMN_USAGE WHERE ' \
            'REFERENCED_TABLE_SCHEMA = (SELECT DATABASE()) AND ' \
            'REFERENCED_TABLE_NAME=%s ORDER BY TABLE_SCHEMA,TABLE_NAME,' \
            'ORDINAL_POSITION'

    conn.execute(query, (table_name,))

    for one_row in conn.fetchall():
        cons_name = one_row[4]
        if cons_name in constraints:
            cur_cons = constraints[cons_name]
        else:
            if verbose:
                logger.info(f'Found constraint {cons_name} on table {one_row[1]}')
            cur_cons = {
                'table_schema': one_row[0],
                'table_name': one_row[1],
                'table_column': one_row[2],
                'constraint_schema': one_row[3],
                'ref_table_schema': one_row[5],
                'ref_table_name': one_row[6],
                'ref_columns':[]
            }

        # Fill in the column information
        ord_pos = int(one_row[8])
        while len(cur_cons['ref_columns']) < ord_pos:
            cur_cons['ref_columns'].append('')
        cur_cons['ref_columns'][ord_pos-1] = one_row[7]

        constraints[cons_name] = cur_cons

    conn.reset()
    if verbose:
        logger.info(f'Found {len(constraints)} constraints referencing table "{table_name}"')
    return constraints if constraints else None


def db_remove_fk_constraints(conn: A2Database, constraints: dict, logger: logging.Logger,
                             verbose: bool=False) -> None:
    """Removes the foreign key constraints from the database
    Arguments:
        conn: the database connection
        constraints: the dictionary of constraints (see db_get_fk_constraints)
        logger: logging instance
        verbose: prints additional information when True (truthy)
    """
    if not constraints:
        return

    for cons_name in constraints.keys():
        cons_info = constraints[cons_name]
        query = f'ALTER TABLE {cons_info["table_name"]} DROP FOREIGN KEY {cons_name}'
        if verbose:
            logger.info(query)
        conn.execute(query)
        conn.reset()


def db_restore_fk_constraints(conn: A2Database, constraints: dict, logger: logging.Logger,
                              verbose: bool=False) -> None:
    """Restores foreign key constraints that were removed
    Arguments:
        conn: the database connection
        constraints: the dictionary of constraints (see db_get_fk_constraints)
        logger: logging instance
        verbose: prints additional information when True (truthy)
    """
    if not constraints:
        return

    for cons_name in constraints.keys():
        cons_info = constraints[cons_name]
        ref_columns = list(cons_info['ref_columns'])
        query = f'ALTER TABLE {cons_info["table_name"]} ADD FOREIGN KEY {cons_name} ' \
                f'({cons_info["table_column"]}) REFERENCES {cons_info["ref_table_name"]} (' + \
                ','.join((one_col for one_col in ref_columns)) + ')'

        if verbose:
            logger.info(query)

        conn.execute(query)
        conn.reset()


def map_col_name(sheet_ids: tuple, col_name: str, col_maps: tuple) -> str:
    """Maps the column name if a sheet and column mapping is found. Returns
       either the mapped column name or the original column name
    Arguments:
        sheet_ids: tuple of sheet IDs to match on
        col_name: column name to match
        col_maps: a tuple of column mapping dicts 
                {'sheet_name': <name>,
                 'sheet_col': <name>,
                 'db_col': <name>
                }
    Return:
        Returns either the mapped column name or the original column name
    """
    for one_map in col_maps:
        if one_map['sheet_name'] in sheet_ids and one_map['sheet_col'] == col_name:
            return one_map['db_col']

    return col_name


def process_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet, conn: A2Database, opts: dict) \
                    -> None:
    """Uploads the data in the worksheet
    Arguments:
        sheet: the worksheet to upload
        conn: the database connection
        opts: additional options
    """
    verbose = 'verbose' in opts and opts['verbose']
    resetting = 'reset' in opts and opts['reset']
    saved_constraints = None

    # Get the table name from the sheet title
    table_name = '_'.join(sheet.title.split('_')[:-1])

    if resetting:
        opts['logger'].info(f'Replacing data in table {table_name} with data from ' \
                            f'tab {sheet.title}')
    else:
        opts['logger'].info(f'Updating table {table_name} from tab {sheet.title}')

    # Get the rows iterator
    rows_iter = sheet.iter_rows()

    # Get the column names
    if 'col_names' in opts and opts['col_names']:
        # User specified column names
        col_names = opts['col_names']
    else:
        # Find the row with the column names
        col_names = []
        # Skip to the row with the names
        cnt = 1
        while cnt < opts['col_names_row']:
            _ = next(rows_iter)
        for one_col in next(rows_iter):
            # Check if we're mapping this name
            col_names.append(map_col_name((sheet.title, table_name), one_col.value,
                                          opts['col_name_map']))

    # Find geometry columns
    geom_col_info, col_alias = conn.get_col_info(table_name, col_names, opts['geometry_epsg'],
                                            colX1=opts['point_col_x'], rowY1=opts['point_col_y'])

    # Check if we're resetting the tables
    if 'reset' in opts and opts['reset']:
        saved_constraints = db_get_fk_constraints(conn, table_name, opts['logger'], verbose)
        db_remove_fk_constraints(conn, saved_constraints, opts['logger'], verbose)
        query = f'TRUNCATE TABLE {table_name}'
        if verbose:
            opts['logger'].info(query)
        conn.execute(query)
        conn.reset()

    # Process the rows
    skipped_rows = 0
    added_updated_rows = 0
    for one_row in rows_iter:
        col_values = tuple(one_cell.value for one_cell in one_row)

        # Skip over missing primary keys
        pk_value = col_values[col_names.index(opts['primary_key'])]
        if pk_value is None:
            opts['logger'].info('Skipping data row with null primary key value: ' \
                  f'row {added_updated_rows + skipped_rows + 1}')
            skipped_rows = skipped_rows + 1
            continue

        # Check for existing data and skip this row if it exists and we're not forcing
        data_exists = conn.check_data_exists_pk(table_name, opts['primary_key'], pk_value,
                                                verbose=verbose)
        if data_exists and not opts['force']:
            skipped_rows = skipped_rows + 1
            continue

        # Check if there are changes to an existing row
        if data_exists:
            data_same = conn.check_data_exists(table_name, col_names, col_values,
                                geom_col_info=geom_col_info,
                                verbose=verbose)
            if data_same:
                opts['logger'].info(f'Skipping unchanged data row: primary key value {pk_value}')
                skipped_rows = skipped_rows + 1
                continue

        added_updated_rows = added_updated_rows + 1
        if geom_col_info and conn.epsg != opts['geometry_epsg']:
            col_values = transform_geom_cols(col_names, col_values, geom_col_info, \
                                             opts['geometry_epsg'], conn.epsg)
        conn.add_update_data(table_name, col_names, col_values, col_alias, geom_col_info, \
                             update=data_exists, \
                             primary_key=opts['primary_key'],
                             verbose=verbose)

    if saved_constraints:
        db_restore_fk_constraints(conn, saved_constraints, opts['logger'], verbose)

    if skipped_rows:
        opts['logger'].info('    Processed ' + str(added_updated_rows + skipped_rows) + \
                            f' rows with {skipped_rows} not updated')
    else:
        opts['logger'].info(f'    Processed {added_updated_rows} rows')


def process_esri_row(conn: A2Database, table_name: str, col_names: tuple, col_values: tuple,
                     opts: dict, verbose: bool=False) -> bool:
    """Handles the processing of one row of EsRI data
    Arguments:
        conn: the database connection
        table_name: the name of the target table
        col_names: the row names
        col_values: the row values
        opts: additional options
        verbose: whether verbose mode is enabled
    Returns:
        Returns True if the data was added or updated, and False if no changes were made
    """
    names = list(col_names)
    values = list(col_values)
    geom_col_info = None
    col_alias = None
    epsg = None

    # Get the X and Y field names
    if 'SHAPE@' in names:
        shape_index = names.index('SHAPE@')
        shape_data = values[shape_index]
        names.append('x')
        names.append('y')
        values.append(shape_data['x'])
        values.append(shape_data['y'])
        epsg = shape_data['spatialReference']['wkid']

        # Remove shape entry from names and values
        names.pop(shape_index)
        values.pop(shape_index)

    # Find geometry columns
    geom_col_info, col_alias = (None, {})
    try:
        geom_col_info, col_alias = conn.get_col_info(table_name, names, opts['geometry_epsg'],
                                            colX1=opts['point_col_x'], rowY1=opts['point_col_y'])
    except ValueError as ex:
        if verbose:
            opts['logger'].warning('Geometry Info', exc_info=True, stack_info=True)
        if not opts['force']:
            opts['logger'].warning('Table geometries are different. Specify --force to ignore this issue')
            raise ex
        else:
            opts['logger'].warning('Table geometries are different. Continuing with processing')

    # Skip over missing primary keys
    pk_value = values[names.index(opts['primary_key'])]
    if pk_value is None:
        opts['logger'].info('Skipping data row with null primary key value')
        return False

    # If the data is not in the table already, add it in
    data_exists = conn.check_data_exists_pk(table_name, opts['primary_key'], pk_value,
                                                                            verbose=verbose)
    if data_exists and not opts['force']:
        return False

    if data_exists:
        if conn.check_data_exists(table_name, names, values, geom_col_info=geom_col_info,
                                                                        verbose=verbose):
            opts['logger'].info(f'Skipping unchanged data row: primary key value {pk_value}')
            return False

    # Adding in the data
    if geom_col_info and epsg and conn.epsg != epsg:
        values = transform_geom_cols(names, values, geom_col_info, epsg, conn.epsg)
    conn.add_update_data(table_name, names, values, col_alias, geom_col_info, \
                         update=data_exists, \
                         primary_key=opts['primary_key'],
                         verbose=verbose)

    return True


def process_esri_data(conn: A2Database, endpoint_url: str, client_id: str, feature_id: str,
                        opts: dict) -> None:
    """Downloads and processes the data from the specified feature
    Arguments:
        conn: the database connection
        endpoint_url: the ESRI endpoint to connect to
        client_id: the ESRI app client ID
        feature_id: the feature ID to download data from
        opts: additional options
    """
    verbose = 'verbose' in opts and opts['verbose']

    opts['logger'].info('Pulling data from ESRI')

    # Connect to ESRI
    gis = GIS(endpoint_url, client_id=client_id)

    # Search for the feature layer
    search_res = gis.content.get(feature_id)

    # Get the feature layer
    feature_layer = None
    if len(search_res) > 0:
        if len(search_res.layers) > 0:
            feature_layer = search_res.layers[0]
    else:
        raise ValueError('Unable to access item with ID {feature_id} at {endpoint_url}')

    skipped_rows = 0
    added_updated_rows = 0

    # Process feature and table data
    if feature_layer:
        # Get the table name
        table_name = feature_layer.properties['name']
        opts['logger'].info(f'Processing ESRI table {table_name}')
        if 'table_name_map' in opts and opts['table_name_map'] and table_name in opts['table_name_map']:
            table_name = opts['table_name_map'][table_name]
            opts['logger'].info(f'    table name mapped for database: {table_name}')

        # Check if we're resetting the tables
        saved_constraints = None
        if 'reset' in opts and opts['reset']:
            saved_constraints = db_get_fk_constraints(conn, table_name, opts['logger'], verbose)
            db_remove_fk_constraints(conn, saved_constraints, opts['logger'], verbose)
            query = f'TRUNCATE TABLE {table_name}'
            if verbose:
                opts['logger'].info(query)
            conn.execute(query)
            conn.reset()

        # Get date indexes
        date_indexes = []
        for field_idx, one_field in enumerate(feature_layer.properties['fields']):
            if one_field['type'] == 'esriFieldTypeDate':
                date_indexes.append(field_idx)

        # Process feature data
        for one_res in feature_layer.query('OBJECTID >= 0'):
            values, names = one_res.as_row
            names = tuple(map_col_name((table_name,), one_name, opts['col_name_map']) \
                                                                        for one_name in names)
            if date_indexes:
                values = tuple(values)
                values = tuple(datetime.utcfromtimestamp(values[cur_idx]/1000.0) if cur_idx in \
                                date_indexes else values[cur_idx] for cur_idx in range(0, len(values)))
            try:
                if process_esri_row(conn, table_name, tuple(names), tuple(values), opts, verbose):
                    added_updated_rows = added_updated_rows + 1
                else:
                    skipped_rows = skipped_rows + 1
            except ValueError as ex:
                # Prepare for an error return
                if saved_constraints:
                    db_restore_fk_constraints(conn, saved_constraints, opts['logger'], verbose)
                raise ex

        if saved_constraints:
            db_restore_fk_constraints(conn, saved_constraints, opts['logger'], verbose)

    # Process data from each of the tables
    for one_table in search_res.tables:
        # Get the table name
        table_name = one_table.properties['name']
        opts['logger'].info(f'Processing next ESRI table {table_name}')
        if 'table_name_map' in opts and opts['table_name_map'] and \
                                                        table_name in opts['table_name_map']:
            table_name = opts['table_name_map'][table_name]
            opts['logger'].info(f'    table name mapped for database: {table_name}')

        # Check if we're resetting the tables
        saved_constraints = None
        if 'reset' in opts and opts['reset']:
            saved_constraints = db_get_fk_constraints(conn, table_name, opts['logger'], verbose)
            db_remove_fk_constraints(conn, saved_constraints, opts['logger'], verbose)
            query = f'TRUNCATE TABLE {table_name}'
            if verbose:
                opts['logger'].info(query)
            conn.execute(query)
            conn.reset()

        # Get date indexes
        date_indexes = []
        for field_idx, one_field in enumerate(one_table.properties['fields']):
            if one_field['type'] == 'esriFieldTypeDate':
                date_indexes.append(field_idx)

        for one_res in one_table.query('OBJECTID >= 0'):
            values, names = one_res.as_row
            names = (map_col_name((table_name,), one_name, opts['col_name_map'])
                                                                    for one_name in names)
            if date_indexes:
                values = tuple(values)
                values = tuple(datetime.utcfromtimestamp(values[cur_idx]/1000.0) if cur_idx in \
                            date_indexes else values[cur_idx] for cur_idx in range(0, len(values)))
            if process_esri_row(conn, table_name, tuple(names), tuple(values), opts, verbose):
                added_updated_rows = added_updated_rows + 1
            else:
                skipped_rows = skipped_rows + 1

        if saved_constraints:
            db_restore_fk_constraints(conn, saved_constraints, opts['logger'], verbose)

    if skipped_rows:
        opts['logger'].info('    Processed ' + str(added_updated_rows + skipped_rows) + \
                            f' rows with {skipped_rows} not updated')
    else:
        opts['logger'].info(f'    Processed {added_updated_rows} rows')


def load_excel_file(filepath: str, opts: dict) -> None:
    """Loads the excel file into the database
    Arguments:
        filepath: the path to the file to load
        opts: additional options
    """
    verbose = 'verbose' in opts and opts['verbose']

    required_opts = ('host', 'database', 'password', 'user')

    # Check the opts parameter
    if opts is None:
        raise ValueError('Missing command line parameters')
    if not all(required in opts for required in required_opts):
        opts['logger'].error('Missing required command line database parameters')
        sys.exit(100)

    # Get the user password if they need to specify one
    if opts['password'] is not False:
        print('Connecting to database...')
        opts['password'] = getpass()

    # MySQL connection
    try:
        db_conn = a2database.connect(
            host=opts["host"],
            database=opts["database"],
            password=opts["password"],
            user=opts["user"],
            logger=opts['logger']
        )
    except mysql.connector.errors.ProgrammingError:
        opts['logger'].error('LoadExcelFile', exc_info=True, stack_info=True)
        opts['logger'].error('Please correct errors and try again')
        sys.exit(101)

    # Set the default database EPSG
    db_conn.epsg = opts['database_epsg']

    # Open the EXCEL file and process each tab
    if filepath:
        workbook = load_workbook(filename=filepath, read_only=True, data_only=True)

        opts['logger'].info(f'Updating using {filepath}')

        for one_sheet in workbook.worksheets:
            process_sheet(one_sheet, db_conn, opts)
    else:
        try:
            process_esri_data(db_conn, opts['esri_endpoint'], opts['esri_client_id'],
                                opts['esri_feature_id'], opts)
        except ValueError as ex:
            if verbose:
                logging.getLogger().error('Value exception caught', exc_info=True, stack_info=True)
                user_opts['logger'].error('Stopping processing due to detected problem')
            else:
                user_opts['logger'].error('Stopping processing due to detected problem. Use the ' \
                                                            '--verbose flag for more information')

    db_conn.commit()


if __name__ == '__main__':
    try:
        excel_filename, user_opts = get_arguments(logging.getLogger())
        user_opts['logger'] = init_logging(user_opts['log_filename'], user_opts['debug'])
        load_excel_file(excel_filename, user_opts)
    except Exception:
        if user_opts and 'logger' in user_opts and user_opts['logger']:
            user_opts['logger'].error('Unhandled exception caught', exc_info=True, stack_info=True)
        else:
            logging.getLogger().error('Unhandled exception caught', exc_info=True, stack_info=True)
        sys.exit(250)
