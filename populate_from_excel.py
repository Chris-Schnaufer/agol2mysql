#!/usr/bin/python3
""" This script movs data from an EXCEL spreadsheet to a MySql database
"""

import os
import argparse
import sys
import logging
from getpass import getpass
import openpyxl
from openpyxl import load_workbook
import mysql.connector

import a2database
from a2database import A2Database

# Check if we have the geometry transformation module
try:
    from osgeo import ogr
    from osgeo import osr
except ModuleNotFoundError:
    GEOM_CAN_TRANSFORM = False
else:
    GEOM_CAN_TRANSFORM = True

# The name of our script
SCRIPT_NAME = os.path.basename(__file__)

# Default host name to connect to the database
DEFAULT_HOST_NAME = 'localhost'

# Default number of heading lines in the EXCEL file
DEFAULT_NUM_HEADER_LINES = 1

# Default row that starts the column names
DEFAULT_COL_NAMES_ROW = 1

# The default primary key database name for the sheet data
DEFAULT_PRIMARY_KEY_NAME = 'UAID'

# The default column name of the schema table names
DEFAULT_SCHEMA_TABLE_NAME_COL='Table Name'

# The default column name of the schema field names
DEFAULT_SCHEMA_FIELD_NAME_COL = 'Field Name'

# The default column name of the schema data type
DEFAULT_SCHEMA_DATA_TYPE_COL = 'Type'

# The default column name of the schema data type length
DEFAULT_SCHEMA_DATA_LEN_COL = 'Size'

# The default column name of the schema data type
DEFAULT_SCHEMA_DESCRIPTION_COL = 'Description'

# Default EPSG code for points
DEFAULT_GEOM_EPSG = 26912

# Default name for logging output
DEFAULT_LOG_FILENAME = 'populate_from_excel.out'

# Argparse-related definitions
# Declare the progam description
ARGPARSE_PROGRAM_DESC = 'Uploads data from an excel spreadsheet to the MySQL database with the ' \
                        'option of creating the schema'
# Epilog to argparse arguments
ARGPARSE_EPILOG = 'Seriously consider backing up your database before running this script to '\
                  'create a schema. Data duplicates are avoided by checking if the data ' \
                  'already exists in the database. Don\'t specify the schema sheet name on ' \
                  'the command line to avoid a table creation - the existing table WILL BE ' \
                  'DROPPED if the --force and --schema_sheet_name options are specified  '
# Host name help
ARGPARSE_HOST_HELP = f'The database host to connect to (default={DEFAULT_HOST_NAME})'
# Name of the database to connect to
ARGPARSE_DATABASE_HELP = 'The database to connect to'
# User name help
ARGPARSE_USER_HELP = 'The username to connect to the database with'
# Password help
ARGPARSE_PASSWORD_HELP = 'The password used to connect to the database (leave empty to be prompted)'
# Declare the help text for the EXCEL filename parameter (for argparse)
ARGPARSE_EXCEL_FILE_HELP = 'Path to the EXCEL file to upload'
# Declare the help text for the force deletion flag
ARGPARSE_UPDATE_HELP = 'Update existing data with new values (default is to skip updates)'
# Help text for the name of the sheet with the data
ARGPARSE_DATA_SHEET_NAME = 'The name of the sheet that contains the data to load'
# Help text for the name of the sheet with the schema
ARGPARSE_SCHEMA_SHEET_NAME = 'The name of the sheet with the schema definition'
# Declare the help for headers
ARGPARSE_HEADER1_HELP = 'Specify the number of lines to consider as headings for the data ' \
                       f'sheet (default {DEFAULT_NUM_HEADER_LINES} lines)'
# Declare the help for where the column names are
ARGPARSE_COL_NAMES1_ROW_HELP = 'Specify the row that contains the column names for the data ' \
                              f'sheet (default row is {DEFAULT_COL_NAMES_ROW})'
# Declare the help for headers
ARGPARSE_HEADER2_HELP = 'Specify the number of lines to consider as headings for the schema ' \
                       f'sheet (default {DEFAULT_NUM_HEADER_LINES} lines)'
# Declare the help for where the column names are
ARGPARSE_COL_NAMES2_ROW_HELP = 'Specify the row that contains the column names for the schema ' \
                              f'sheet (default row is {DEFAULT_COL_NAMES_ROW})'
# The help to indicate that only the schema is to be created and no data loaded
ARGPARSE_SCHEMA_ONLY = 'Specify this flag if only the schema is to be created or updated and no ' \
                       'data loaded'
# Sheet column containing the table field names in the schema sheet
ARGPARSE_SCHEMA_TABLE_NAME_COL_HELP = 'The column name containing the schema table names ' \
                            f'(default name is "{DEFAULT_SCHEMA_TABLE_NAME_COL}"). You ' \
                            'can also specify the numerical index of the column (starting at 1)'
# Sheet column containing the table column names in the schema sheet
ARGPARSE_SCHEMA_FIELD_NAME_COL_HELP = 'The column name containing the schema field names ' \
                            f'(default name is "{DEFAULT_SCHEMA_FIELD_NAME_COL}"). You ' \
                            'can also specify the numerical index of the column (starting at 1)'
# Sheet column containing the column data types in the schema sheet
ARGPARSE_SCHEMA_DATA_TYPE_COL_HELP = 'The column name containing the schema data types (default ' \
                            f'name is "{DEFAULT_SCHEMA_DATA_TYPE_COL}"). You can also specify ' \
                            'the numerical index of the column (starting at 1)'
# Sheet column containing the column data type lengths in the schema sheet
ARGPARSE_SCHEMA_DATA_LEN_COL_HELP = 'The column name containing the schema data type lengths ' \
                            f'(default name is "{DEFAULT_SCHEMA_DATA_LEN_COL}"). You can also ' \
                            'specify the numerical index of the column (starting at 1)'
# Sheet column containing the column description in the schema sheet
ARGPARSE_SCHEMA_DESCRIPTION_COL_HELP = 'The column name containing the schema description column ' \
                            f'(default name is "{DEFAULT_SCHEMA_DESCRIPTION_COL}"). You can also ' \
                            'specify the numerical index of the column (starting at 1)'
# Help text for using all defined columns in the schema sheet when creating the table
ARGPARSE_USE_SCHEMA_COLS_HELP = 'Specify if all the defined columns in the schema sheet should ' \
                            'be used when creating the schema. Defaults to only creating the ' \
                            'columns that are found in the data sheet'
# Help text fo specifying the primary key to use
ARGPARSE_PRIMARY_KEY_HELP = 'The name of the primary key to use when checking if data is ' \
                            f'already in the database (default "{DEFAULT_PRIMARY_KEY_NAME}"). ' \
                            'Also needs to be a column in the spreadsheet'
# Help for specifying columns that are to be considered point columns
ARGPARSE_POINT_COLS_HELP = 'The names of the X and Y columns in the spreadsheet that represent ' \
                           'a point type when creating a schema ("<x name>,<y name>")'
# Help for specifying the EPSG code that the point coordinates are in
ARGPARSE_GEOMETRY_EPSG_HELP = 'The EPSG code of the coordinate system for the geometric values ' \
                           f'(default is {DEFAULT_GEOM_EPSG})'
# Help for specifying the default database connection EPSG code
ARGPARSE_DATABASE_EPSG_HELP = 'The EPSG code of the database geometry ' \
                           f'(default is {DEFAULT_GEOM_EPSG})'
# Help for supressing the creation of views
ARGPARSE_NOVIEWS_HELP = 'For tables created with geometry, do not create an associated ' \
                        'view that breaks out the different values. Associated views are ' \
                        'created by default when a geometry is present'
# Help text for verbose flag
ARGPARSE_VERBOSE_HELP = 'Display additional information as the script is run'
# Help text for ignore column parameters
ARGPARSE_IGNORE_COL_HELP = 'A column to ignore when loading the data. You can also specify the ' \
                           'numerical index of the column (starting at 1). Can use multiple ' \
                           'times on the command line'
ARGPARSE_FORCE_PK_TEXT_HELP = 'Force the primary to be text (instead of the default INT type)'
# Help for changing the logging filename
ARGPARSE_LOG_FILENAME_HELP = 'Change the name of the file where logging gets saved. This is a ' \
                             'destructive overwrite of existing files. Default logging file is ' \
                             f'named {DEFAULT_LOG_FILENAME}'

def get_arguments(logger: logging.Logger) -> tuple:
    """ Returns the data from the parsed command line arguments
    Returns:
        A tuple consisting of a string containing the EXCEL file name to process, and
        a dict of the command line options
    Exceptions:
        A ValueError exception is raised if the filename is not specified
    Notes:
        If an error is found, the script will exit with a non-zero return code
    """
    parser = argparse.ArgumentParser(prog=SCRIPT_NAME,
                                     description=ARGPARSE_PROGRAM_DESC,
                                     epilog=ARGPARSE_EPILOG)
    parser.add_argument('excel_file', nargs='+', help=ARGPARSE_EXCEL_FILE_HELP)
    parser.add_argument('-o', '--host', default=DEFAULT_HOST_NAME, help=ARGPARSE_HOST_HELP)
    parser.add_argument('-d', '--database', help=ARGPARSE_DATABASE_HELP)
    parser.add_argument('-u', '--user', help=ARGPARSE_USER_HELP)
    parser.add_argument('-f', '--force', action='store_true', help=ARGPARSE_UPDATE_HELP)
    parser.add_argument('--verbose', action='store_true', help=ARGPARSE_VERBOSE_HELP)
    parser.add_argument('-p', '--password', action='store_true', help=ARGPARSE_PASSWORD_HELP)
    parser.add_argument('-dt', '--data_sheet_name', help=ARGPARSE_DATA_SHEET_NAME)
    parser.add_argument('-dh', '--data_header', type=int, default=DEFAULT_NUM_HEADER_LINES,
                        help=ARGPARSE_HEADER1_HELP)
    parser.add_argument('-dn', '--data_col_names_row', type=int,
                        default=DEFAULT_COL_NAMES_ROW,
                        help=ARGPARSE_COL_NAMES1_ROW_HELP)
    parser.add_argument('--schema_only', action='store_true', help=ARGPARSE_SCHEMA_ONLY)
    parser.add_argument('-st', '--schema_sheet_name', help=ARGPARSE_SCHEMA_SHEET_NAME)
    parser.add_argument('-sh', '--schema_header', type=int, default=DEFAULT_NUM_HEADER_LINES,
                        help=ARGPARSE_HEADER2_HELP)
    parser.add_argument('-sc', '--schema_col_names_row', type=int, default=DEFAULT_COL_NAMES_ROW,
                        help=ARGPARSE_COL_NAMES2_ROW_HELP)
    parser.add_argument('-stn', '--schema_table_name_col', default=DEFAULT_SCHEMA_TABLE_NAME_COL,
                        help=ARGPARSE_SCHEMA_TABLE_NAME_COL_HELP)
    parser.add_argument('-sfn', '--schema_field_name_col', default=DEFAULT_SCHEMA_FIELD_NAME_COL,
                        help=ARGPARSE_SCHEMA_FIELD_NAME_COL_HELP)
    parser.add_argument('-sft', '--schema_data_type_col', default=DEFAULT_SCHEMA_DATA_TYPE_COL,
                        help=ARGPARSE_SCHEMA_DATA_TYPE_COL_HELP)
    parser.add_argument('-sfl', '--schema_data_len_col', default=DEFAULT_SCHEMA_DATA_LEN_COL,
                        help=ARGPARSE_SCHEMA_DATA_LEN_COL_HELP)
    parser.add_argument('-sfd', '--schema_description_col',
                        default=DEFAULT_SCHEMA_DESCRIPTION_COL,
                        help=ARGPARSE_SCHEMA_DESCRIPTION_COL_HELP)
    parser.add_argument('--use_schema_cols', action='store_true',
                        help=ARGPARSE_USE_SCHEMA_COLS_HELP)
    parser.add_argument('--point_cols', help=ARGPARSE_POINT_COLS_HELP)
    parser.add_argument('--geometry_epsg', type=int, default=DEFAULT_GEOM_EPSG,
                        help=ARGPARSE_GEOMETRY_EPSG_HELP)
    parser.add_argument('--database_epsg', type=int, default=DEFAULT_GEOM_EPSG,
                        help=ARGPARSE_DATABASE_EPSG_HELP)
    parser.add_argument('-k', '--key_name', default=DEFAULT_PRIMARY_KEY_NAME,
                        help=ARGPARSE_PRIMARY_KEY_HELP)
    parser.add_argument('--noviews', action='store_true',
                        help=ARGPARSE_NOVIEWS_HELP)
    parser.add_argument('--ignore_col', nargs='*', action='extend', help=ARGPARSE_IGNORE_COL_HELP)
    parser.add_argument('--pk_force_text', action='store_true', help=ARGPARSE_FORCE_PK_TEXT_HELP)
    parser.add_argument('--log_filename', default=DEFAULT_LOG_FILENAME,
                        help=ARGPARSE_LOG_FILENAME_HELP)
    args = parser.parse_args()

    # Find the EXCEL file and the password (which is allowed to be eliminated)
    excel_file = None
    if not args.excel_file:
        # Raise argument error
        raise ValueError('Missing a required argument')

    if len(args.excel_file) == 1:
        excel_file = args.excel_file[0]
    else:
        # Report the problem
        logger.error('Too many arguments specified for input file')
        parser.print_help()
        sys.exit(10)

    # Check that we can access the EXCEL file
    try:
        with open(excel_file, encoding="utf-8"):
            pass
    except FileNotFoundError:
        logger.error(f'Unable to open EXCEL file {excel_file}')
        sys.exit(11)

    # Check point column names parameter
    if args.point_cols:
        # Point columns come in pairs
        if not ',' in args.point_cols:
            logger.error('Point column names must be separated by a comma (,)')
            sys.exit(12)
        # Checking for non-blank names
        for one_name in (one_col.strip() for one_col in args.point_cols.split(',')):
            if not one_name:
                logger.error('Please specify an X and Y column name for point support')
                sys.exit(13)

    cmd_opts = {'force': args.force,
                'verbose': args.verbose,
                'host': args.host,
                'database': args.database,
                'user': args.user,
                'password': args.password ,
                'data_sheet_name': args.data_sheet_name,
                'data_header_lines': args.data_header,
                'data_col_names_row': args.data_col_names_row,
                'schema_only': args.schema_only,
                'schema_sheet_name': args.schema_sheet_name,
                'schema_header_lines': args.schema_header,
                'schema_col_names_row': args.schema_col_names_row,
                'schema_table_name_col': args.schema_table_name_col,
                'schema_field_name_col': args.schema_field_name_col,
                'schema_data_type_col': args.schema_data_type_col,
                'schema_data_len_col': args.schema_data_len_col,
                'schema_description_col': args.schema_description_col,
                'use_schema_cols': args.use_schema_cols,
                'ignore_cols': tuple(one_col.casefold() for one_col in args.ignore_col),
                'point_col_x': args.point_cols.split(',')[0] if args.point_cols else None,
                'point_col_y': args.point_cols.split(',')[1] if args.point_cols else None,
                'geometry_epsg': args.geometry_epsg,
                'database_epsg': args.database_epsg,
                'primary_key': args.key_name,
                'primary_key_text': args.pk_force_text,
                'noviews': args.noviews,
                'log_filename': args.log_filename
               }

    # Return the loaded JSON
    return excel_file, cmd_opts


def init_logging(filename: str) -> logging.Logger:
    """Initializes the logging
    Arguments:
        filename: name of the file to save logging to
    Return:
        Returns the created logger instance
    """
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    # Create formatter
    formatter = logging.Formatter('%(levelname)s: %(message)s')

    # Console handler
    cur_handler = logging.StreamHandler()
    cur_handler.setFormatter(formatter)
    logger.addHandler(cur_handler)

    # Output file handler
    cur_handler = logging.FileHandler(filename, mode='w')
    cur_handler.setFormatter(formatter)
    logger.addHandler(cur_handler)

    return logger


def transform_points(from_epsg: int, to_epsg: int, values: tuple) -> list:
    """Returns the list with the last num_values converted to the new coordinate system
    Arguments:
        from_epsg: original EPSG code
        to_epsg: EPSG code to transform the points to
        values: the list of values to transform (assumes X1, Y1, X2, Y2, ...)
    Returns:
        The list of transformed values
    """
    if len(values) % 2 or len(values) < 1:
        raise ValueError('Unable to transform points, an even number of X,Y values pairs ' \
                         'are needed')

    # Make sure we can transform points
    if not GEOM_CAN_TRANSFORM:
        raise ValueError('Unable to transform points, supporting osgeo module is not installed')

    # Transform the points
    return_values = []
    idx = 0
    from_sr = osr.SpatialReference()
    from_sr.ImportFromEPSG(int(from_epsg))
    to_sr = osr.SpatialReference()
    to_sr.ImportFromEPSG(int(to_epsg))
    transform = osr.CreateCoordinateTransformation(from_sr, to_sr)
    while idx < len(values):
        cur_geom = ogr.CreateGeometryFromWkt(f'POINT({values[idx]} {values[idx+1]} {from_epsg})')
        cur_geom.Transform(transform)
        return_values.extend((cur_geom.GetX(), cur_geom.GetY()))
        idx = idx + 2

    return return_values


def transform_geom_cols(col_names: tuple, col_values: tuple, geom_col_info: dict, from_epsg: int, \
                        to_epsg: int) -> list:
    """Transforms geometry point to the specified coordinate system
    Arguments:
        col_name: the column names of the table
        col_values: the values associated with the column names
        geom_col_info: the geometry column information
        from_epsg: the EPSG code to tranform from
        to_epsg: the EPSG code to transform to
    Return:
        Returns a list of column values with the geometry values transformed
    """
    # Check if we can avoid the transformations
    if from_epsg == to_epsg:
        return list(col_values)
    if len(geom_col_info['sheet_cols']) % 2:
        raise ValueError('Invalid number of X,Y column name pairs specified')

    pt_indexes = []
    pt_values = []
    idx = 0
    while idx < len(geom_col_info['sheet_cols']):
        x_idx = col_names.index(geom_col_info['sheet_cols'][idx])
        y_idx = col_names.index(geom_col_info['sheet_cols'][idx+1])
        pt_indexes.append(x_idx)
        pt_indexes.append(y_idx)
        pt_values.append(col_values[x_idx])
        pt_values.append(col_values[y_idx])
        idx += 2

    new_pt_values = transform_points(from_epsg, to_epsg, pt_values)

    return_values = list(col_values)
    for idx, idx_val in enumerate(pt_indexes):
        return_values[idx_val] = new_pt_values[idx]

    return return_values


def map_col_type(col_type: str, col_len: int=None, raise_on_error: bool=False) -> str:
    """Maps the column type to a MySQL type
    Arguments:
        col_type: the string representing the column type
        col_len: Optional length specification for column (may not be honored)
        raise_on_error: When True, raise an IndexError if the column type is unknown
    Returns:
        Returns the mapped string type
    Exceptions:
        Raises an IndexError if the type is not known
    """
    col_ret_type = None
    match col_type:
        case 'Number':
            col_ret_type = 'DOUBLE'

        case 'Short Text' | 'Long Text':
            if not col_len:
                col_ret_type = 'VARCHAR(512)'
            else:
                col_ret_type = f'VARCHAR({col_len})'

        case 'Date/Time':
            col_ret_type = 'TIMESTAMP'

        case 'Yes/No':
            col_ret_type = 'TINYINT'

        case 'POINT':
            col_ret_type = 'POINT'

    if col_ret_type is None and raise_on_error:
        raise IndexError(f'Unknown column type {col_type} found')

    return col_ret_type


def db_update_schema(table_name: str, schema_sheet: openpyxl.worksheet.worksheet.Worksheet, \
                     col_names: tuple, opts: dict, conn: A2Database) -> None:
    """Updates the database schema
    Arguments:
        table_name: the name of the table
        schema_sheet: the sheet containing the schema information
        col_name: the column names from the data sheet
        opts: additional options
        cursor: the database cursor
        conn: the database connection
    """
    # Define some handy variables
    lower_col_names = tuple((one_name.casefold() for one_name in col_names))
    verbose = 'verbose' in opts and opts['verbose']

    # Check if the table exists
    table_exists = conn.table_exists(table_name)
    if table_exists:
        if 'force' not in opts or not opts['force']:
            if verbose:
                opts['logger'].warning(f'Table {table_name} already exists and the ' \
                                       'force flag is not specified')
                opts['logger'].warning('    not updating table')
            return

    # If we have point columns specified, check that they are valid
    point_col_names = None
    if opts["point_col_x"]:
        if not opts["point_col_x"].casefold() in lower_col_names:
            raise ValueError(f'The X column name for point is not found "{opts["point_col_x"]}"')
        if not opts["point_col_y"].casefold() in lower_col_names:
            raise ValueError(f'The Y column name for point is not found "{opts["point_col_y"]}"')
        point_col_names = (opts["point_col_x"].casefold(), opts["point_col_y"].casefold())

    # Load all the indexes into the schema definition sheet
    col_table_idx = int(opts['schema_table_name_col']) - 1 \
                        if opts['schema_table_name_col'].isnumeric() else None
    col_name_idx = int(opts['schema_field_name_col']) - 1 \
                        if opts['schema_field_name_col'].isnumeric() else None
    col_type_idx = int(opts['schema_data_type_col']) - 1 \
                        if opts['schema_data_type_col'].isnumeric() else None
    col_len_idx = int(opts['schema_data_len_col']) - 1 \
                        if opts['schema_data_len_col'].isnumeric() else None
    col_desc_idx = int(opts['schema_description_col']) - 1 \
                        if opts['schema_description_col'].isnumeric() else None

    # Get the rows iterator
    rows_iter = schema_sheet.iter_rows()

    # Get the column names and find the indexes we're looking for
    if opts['schema_col_names_row'] > 0:
        # Skip to the row with the names
        for _ in range(1, opts['schema_col_names_row']):
            _ = next(rows_iter)
        idx = 0
        for one_col in next(rows_iter):
            if one_col.value is None:
                continue
            cur_name = one_col.value.casefold()
            if cur_name == opts['schema_table_name_col'].casefold():
                col_table_idx = idx
            elif cur_name == opts['schema_field_name_col'].casefold():
                col_name_idx = idx
            elif cur_name == opts['schema_data_type_col'].casefold():
                col_type_idx = idx
            elif cur_name == opts['schema_data_len_col'].casefold():
                col_len_idx = idx
            elif cur_name == opts['schema_description_col'].casefold():
                col_desc_idx = idx
            idx = idx + 1

    if col_name_idx is None or col_type_idx is None or col_desc_idx is None:
        raise IndexError('Unable to find schema columns')

    # Skip over header lines
    if opts['schema_col_names_row'] < opts['schema_header_lines']:
        for _ in range(opts['schema_col_names_row'], opts['schema_header_lines']):
            _ = next(rows_iter)

    # Prepare the column information
    col_info = []
    ignore_columns = opts['ignore_cols'] if 'ignore_cols' in opts else []
    for one_row in rows_iter:
        # Skip if we're only adding columns found in the data sheet and it's not a match
        if 'use_schema_cols' not in opts or not opts['use_schema_cols']:
            if one_row[col_table_idx].value is None:
                continue
            if one_row[col_table_idx].value.casefold() != table_name.casefold():
                continue
            if one_row[col_name_idx].value is None:
                continue
            if one_row[col_name_idx].value.casefold() not in lower_col_names:
                continue
        # Skip over the point column names if we're creating a point column
        if point_col_names and one_row[col_name_idx].value.casefold() in point_col_names:
            continue
        # Make sure this column belongs to the current table
        col_table = one_row[col_table_idx].value
        if col_table.casefold() != table_name.casefold():
            continue
        # Check if we ignore a column
        col_name = one_row[col_name_idx].value
        if col_name.casefold() in ignore_columns:
            continue
        # Add the column information to the list
        col_type = map_col_type(one_row[col_type_idx].value,
                        int(one_row[col_len_idx].value) if one_row[col_len_idx].value else 0,
                        raise_on_error=True)
        is_primary = col_name.casefold() == opts['primary_key'].casefold()
        is_primary_text = 'primary_key_text' in opts and opts['primary_key_text']
        col_info.append({
            'name': col_name,
            'type': 'INT' if is_primary and not is_primary_text else col_type,
            'is_primary': is_primary,
            'auto_increment': is_primary and not is_primary_text, # Primary key auto-increment
            'description': one_row[col_desc_idx].value,
            'null_allowed': not is_primary,
            'index': is_primary                 # Primary keys are indexed
            })

    # Make sure we have something
    if not col_info:
        opts['logger'].warning(f'No column descriptions found for table {table_name}')
        return

    # Add in the point column type if we're creating one
    if point_col_names:
        col_info.append({
            'name': 'geom',
            'type': 'POINT',
            'is_spatial': True,
            'description': 'Auto-generated column',
            'null_allowed': False,
            'index': True
            })

    # If the table exists, we need to drop it
    if table_exists:
        conn.drop_table(table_name, verbose)

    # Create the table
    conn.create_table(table_name, col_info, verbose)

    # Create a view if we have grometries
    if point_col_names:
        view_name = table_name + '_view'
        if ('noviews' not in opts or not opts['noviews']):
            conn.create_view(view_name, table_name, col_info, verbose)
        elif 'force' in opts and opts['force']:
            conn.drop_view(view_name)


def process_sheets(data_sheet: openpyxl.worksheet.worksheet.Worksheet, \
                   schema_sheet: openpyxl.worksheet.worksheet.Worksheet, \
                   opts: dict, conn: A2Database) -> None:
    """Uploads the data in the worksheet
    Arguments:
        data_sheet: the worksheet with data to upload
        schema_sheet: the worksheeet with the table schema information
        opts: additional options
        conn: the database connection
    """
    # Get the table name from the sheet title
    table_name = data_sheet.title
    ignore_columns = tuple(one_ignore.casefold() for one_ignore in opts['ignore_cols']) \
                                                    if 'ignore_cols' in opts else tuple()

    opts['logger'].info(f'Updating table {table_name} from sheet {data_sheet.title}')

    # Get the rows iterator
    rows_iter = data_sheet.iter_rows()

    # Get the column names
    col_names = []
    # Skip to the row with the names
    cnt = 1
    while cnt < opts['data_col_names_row']:
        _ = next(rows_iter)
        cnt = cnt + 1

    # Get the column names
    ignore_idx = []
    idx = 0
    for one_col in next(rows_iter):
        if one_col.value is not None and one_col.value.casefold() not in ignore_columns:
            col_names.append(one_col.value)
        else:
            ignore_idx.append(idx)
        idx += 1

    # Add/Change the schema
    if schema_sheet:
        db_update_schema(table_name, schema_sheet, col_names, opts, conn)
        opts['logger'].info(f'    Updated the schema for {table_name}')
        if opts["schema_only"]:
            return

    # Find geometry columns
    geom_col_info, col_alias = conn.get_col_info(table_name, col_names, opts['geometry_epsg'],
                                            colX1=opts['point_col_x'], rowY1=opts['point_col_y'])

    # Process the rows
    skipped_rows = 0
    added_updated_rows = 0
    for one_row in rows_iter:
        col_values = tuple(one_cell.value for idx,one_cell in enumerate(one_row) \
                                                                        if idx not in ignore_idx)
        pk_value = col_values[col_names.index(opts['primary_key'])]
        if pk_value is None:
            opts['logger'].info('Skipping row with null primary key value: ' \
                  f'row {added_updated_rows + skipped_rows + opts["data_col_names_row"] + 1}')
            skipped_rows = skipped_rows + 1
            continue

        # Check for existing data and skip this row if it exists and we're not forcing
        data_exists = conn.check_data_exists(table_name, col_names, col_values,
                                            geom_col_info=geom_col_info,
                                            primary_key=opts['primary_key'],
                                            verbose=opts['verbose'] if 'verbose' in opts else False)
        if data_exists and not opts['force']:
            skipped_rows = skipped_rows + 1
            continue

        added_updated_rows = added_updated_rows + 1
        if geom_col_info and conn.epsg != opts['geometry_epsg']:
            col_values = transform_geom_cols(col_names, col_values, geom_col_info, \
                                             opts['geometry_epsg'], conn.epsg)
        conn.add_update_data(table_name, col_names, col_values, col_alias, geom_col_info, \
                             update=data_exists, \
                             primary_key=opts['primary_key'], verbose=opts['verbose'])

    if skipped_rows:
        opts['logger'].info(f'    Processed {added_updated_rows + skipped_rows}' \
                            f'data rows with {skipped_rows} not updated')
    else:
        opts['logger'].info(f'    Processed {added_updated_rows + skipped_rows} data rows')


def confirm_options(opts: dict, workbook: openpyxl.workbook.workbook.Workbook) -> tuple:
    """Checks the options specified against the workbook
    Arguments:
        opts: the options to check
        workbook: the workbook to check against
    Returns:
        Returns the data and the (optional) schema sheets
    Notes:
        Will print out any problems to stdout
    """
    # Check if we're schema only
    schema_only = opts['schema_only'] if 'schema_only' in opts else False

    # Confirm the sheets exist
    if not 'data_sheet_name' in opts:
        opts['logger'].error('You need to specify the data sheet name')
        return None, None
    if not opts['data_sheet_name'] or opts['data_sheet_name'] not in workbook.sheetnames:
        opts['logger'].error(f'Unable to find sheet {opts["data_sheet_name"]} in excel file')
        return None, None
    if 'schema_sheet_name' in opts and opts['schema_sheet_name']:
        if opts['schema_sheet_name'] not in workbook.sheetnames:
            opts['logger'].error(f'Unable to find schema sheet {opts["schema_sheet_name"]} ' \
                                  'in excel file')
            return None, None
    if schema_only and not opts['schema_sheet_name']:
        opts['logger'].warning('Schema only is set but no schema sheet name is specified')

    # We don't check column information here

    return workbook[opts['data_sheet_name']], \
           workbook[opts['schema_sheet_name']] if 'schema_sheet_name' in opts and \
                                                    opts['schema_sheet_name'] else None


def load_excel_file(filepath: str, opts: dict) -> None:
    """Loads the excel file into the database
    Arguments:
        filepath: the path to the file to load
        opts: additional options
    """
    required_opts = ('host', 'database', 'password', 'user', 'data_sheet_name')

    # Check the opts parameter
    if opts is None:
        raise ValueError('Missing command line parameters')
    if not all(required in opts for required in required_opts):
        opts['logger'].error('Missing required command line database parameters')
        sys.exit(100)

    # Get the user password if they need to specify one
    if opts['password'] is not False:
        opts['password'] = getpass()

    # MySQL connection
    try:
        db_conn = a2database.connect(
            host=opts["host"],
            database=opts["database"],
            password=opts["password"],
            user=opts["user"],
            logger=opts['logger']
        )
    except mysql.connector.errors.ProgrammingError:
        opts['logger'].error('', exc_info=True)
        opts['logger'].error('Please correct errors and try again')
        sys.exit(101)

    # Set the default database EPSG
    db_conn.epsg = opts['database_epsg']

    # Open the EXCEL file
    workbook = load_workbook(filename=filepath, read_only=True, data_only=True)

    # Make sure the values specified as parameters make sense
    data_sheet, schema_sheet = confirm_options(opts, workbook)
    if not data_sheet:
        sys.exit(102)

    opts['logger'].info(f'Updating using {filepath}')

    process_sheets(data_sheet, schema_sheet, opts, db_conn)

    db_conn.commit()


if __name__ == '__main__':
    try:
        excel_filename, user_opts = get_arguments(logging.getLogger())
        user_opts['logger'] = init_logging(user_opts['log_filename'])
        load_excel_file(excel_filename, user_opts)
    except Exception:
        user_opts['logger'].error('Unhandled exception caught', exc_info=True, stack_info=True)
        sys.exit(250)
